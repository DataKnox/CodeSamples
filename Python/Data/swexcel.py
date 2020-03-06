import openpyxl

wb = openpyxl.load_workbook('/home/knox/Documents/starwars.xlsx')
ws = wb.active

row_count = ws.max_row+1

characters = []

for x in range(2, row_count):
    character = {}
    character['first_name'] = ws.cell(row=x, column=1).value
    character['last_name'] = ws.cell(row=x, column=2).value
    character['homeplant'] = ws.cell(row=x, column=3).value
    characters.append(character)

for char in characters:
    print(char)
